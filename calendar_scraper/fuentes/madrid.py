"""Fuente: Federación de Atletismo de Madrid (atletismomadrid.com).

La FAM publica el calendario MENSUAL en artículos Joomla distintos (un id de
artículo por mes), por lo que scrapear un único artículo (antes id=3292) solo
traía ese mes y se perdían julio, agosto… En cambio expone la TEMPORADA COMPLETA
como Excel. Leemos ese Excel (todos los meses de una vez) con openpyxl. Columnas:
  Fecha · Día · Fecha Fin · Día · Competición · Lugar · Tipo · Últ. modificación.

OJO (ago 2026): la FAM retiró el export antiguo (`mod_calendario/excel/...php`,
ahora 404) y lo sustituyó por un endpoint AJAX de Joomla (`com_ajax`, parámetro
`season`). Probamos las variantes por orden y usamos la primera que devuelva un
.xlsx de verdad (firma PK). El 404 antiguo estuvo SEMANAS tragado por un
`except: continue`; hoy los errores se propagan y quedan en el resumen — esa es
la política de fuentes/__init__.py, adaptada del motor fantasy-atletismo-engine.

Va tras Cloudflare; `http.Http` (curl_cffi impersonate chrome) lo pasa en local.
"""

from __future__ import annotations

import io
from datetime import date, datetime

from calendar_scraper.http import Http, HttpError
from calendar_scraper.models import Competicion, Inscripcion

BASE = "https://www.atletismomadrid.com"
# Página de calendario (vista web del mes en curso): la usamos como enlace de ficha.
CAL = f"{BASE}/index.php?option=com_content&view=article&id=3292&Itemid=111"
# Variantes del export de temporada, en orden de preferencia:
#   1. endpoint com_ajax actual (visto en el href "Exportar" de la página id=3292)
#   2. export legado (por si la FAM lo restaura)
EXPORTS = (
    BASE + "/component/ajax/?module=calendario&method=export&format=raw"
           "&module_id=205&season={anio}&Itemid=111",
    BASE + "/modules/mod_calendario/excel/calendario_excel.php?temporada={anio}&Itemid=111",
)

_DISC = {
    "AL": "Pista Aire libre", "PC": "Pista Cubierta", "R": "Ruta",
    "M": "Marcha", "TR-MT": "Trail", "TR": "Trail", "CT": "Cross", "O": "Cross",
}


def _a_fecha(v: object) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _texto(v: object) -> str | None:
    s = str(v).strip() if v is not None else ""
    return s or None


def _descargar_excel(http: Http, anio: int) -> bytes:
    """Prueba las variantes del export y devuelve el primer .xlsx real (firma PK).

    Si TODAS fallan, propaga el último error: eso es un fallo de la fuente, no
    "temporada sin competiciones", y debe constar en el resumen.
    """
    ultimo: Exception | None = None
    for plantilla in EXPORTS:
        try:
            data = http.get_bytes(plantilla.format(anio=anio))
            if data[:2] == b"PK":
                return data
            ultimo = RuntimeError(
                f"el export no devolvió un xlsx (¿HTML de error?): {plantilla.format(anio=anio)}"
            )
        except HttpError as exc:
            ultimo = exc
    raise ultimo if ultimo else RuntimeError("sin variantes de export configuradas")


def listar(http: Http, desde: date, hasta: date) -> list[Competicion]:
    # Los errores se PROPAGAN a propósito (ver política en fuentes/__init__.py).
    # Madrid es justo el caso que motivó la regla: va tras Cloudflare y desde la IP
    # de CI el Excel puede devolver 403. Tragarse eso y seguir con `out` vacío hacía
    # que el calendario pareciese correcto (conserva las competiciones del JSON
    # anterior) mientras Madrid llevaba semanas sin actualizarse.
    from openpyxl import load_workbook

    out: list[Competicion] = []
    for anio in range(desde.year, hasta.year + 1):
        data = _descargar_excel(http, anio)
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 6:
                    continue
                ini = _a_fecha(row[0])
                if not ini or not (desde <= ini <= hasta):
                    continue
                nombre = _texto(row[4])
                if not nombre:
                    continue
                fin = _a_fecha(row[2]) or ini
                if fin < ini:
                    fin = ini
                tipo = (str(row[6]).strip().upper() if len(row) > 6 and row[6] else "")
                out.append(Competicion(
                    nombre=nombre,
                    ambito="Madrid",
                    disciplina=_DISC.get(tipo),
                    fecha_inicio=ini,
                    fecha_fin=fin,
                    lugar=_texto(row[5]),
                    comunidad="Madrid",
                    pruebas=[],
                    inscripcion=Inscripcion(
                        instrucciones="Inscripción y normas en la web de la Federación de Madrid.",
                    ),
                    url_detalle=CAL,
                    url_calendario_federacion=CAL,
                    fuente="madrid",
                ))
        finally:
            wb.close()
    return out
